import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
import sys


def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


EXCEL_FILE = os.path.join(get_base_dir(), "detailing_records.xlsx")

SERVICES = [
    "Полировка",
    "Керамика",
    "Химчистка",
    "Полимер",
    "Оклейка",
    "Покраска дисков",
    "Покраска кузова",
    "ПДР",
    "Мойка мотора",
    "Реставрация кожи",
    "Тонировка",
    "Мойка кузова",
    "Мойка арок",
    "Реставрация фар"
]

HEADERS = [
    "#", "Марка", "г.номер", "Владелец",
    "Услуга", "Цена", "Материал",
    "Сумма рабочего", "Расход", "Касса", "Примечание"
]


class DetailingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Детейлинг — Учёт клиентов")
        self.root.minsize(460, 520)
        self.root.resizable(True, True)
        self.root.configure(bg="#f5f5f5")

        self._build_ui()

    def _build_ui(self):
        # Header
        header = tk.Frame(self.root, bg="#1a1a2e", pady=12)
        header.pack(fill="x")
        tk.Label(
            header, text="🚗  Детейлинг — Учёт клиентов",
            font=("Helvetica", 15, "bold"),
            fg="white", bg="#1a1a2e"
        ).pack()

        # Main form frame
        form_frame = tk.Frame(self.root, bg="#f5f5f5", padx=24, pady=16)
        form_frame.pack(fill="both", expand=True)

        self.entries = {}

        def add_label_entry(row, label, key, readonly=False, hint=""):
            tk.Label(
                form_frame, text=label, anchor="w",
                font=("Helvetica", 11), bg="#f5f5f5", fg="#333"
            ).grid(row=row, column=0, sticky="w", pady=5, padx=(0, 12))

            if readonly:
                var = tk.StringVar(value="0")
                widget = tk.Label(
                    form_frame, textvariable=var,
                    font=("Helvetica", 11, "bold"),
                    bg="#e8f4e8", fg="#2d6a2d",
                    relief="groove", width=22, anchor="w", padx=6
                )
                widget.grid(row=row, column=1, sticky="ew", pady=5)
                self.entries[key] = var
            else:
                var = tk.StringVar()
                widget = tk.Entry(
                    form_frame, textvariable=var,
                    font=("Helvetica", 11),
                    relief="solid", bd=1, bg="white"
                )
                widget.grid(row=row, column=1, sticky="ew", pady=5)
                if hint:
                    widget.insert(0, hint)
                    widget.config(fg="grey")
                    widget.bind("<FocusIn>", lambda e, w=widget, h=hint: self._clear_hint(e, w, h))
                    widget.bind("<FocusOut>", lambda e, w=widget, h=hint: self._restore_hint(e, w, h))
                self.entries[key] = var

        form_frame.columnconfigure(1, weight=1)

        row_idx = 0
        add_label_entry(row_idx,  "Марка *",       "marka"); row_idx += 1
        add_label_entry(row_idx,  "г.номер *",     "gnomer"); row_idx += 1
        add_label_entry(row_idx,  "Владелец",      "owner",    hint="87XXXXXXXXX"); row_idx += 1

        # Services block
        tk.Label(
            form_frame, text="Услуги и цены *", anchor="w",
            font=("Helvetica", 11, "bold"), bg="#f5f5f5", fg="#333"
        ).grid(row=row_idx, column=0, columnspan=2, sticky="w", pady=(10, 0)); row_idx += 1

        self.services_container = tk.Frame(form_frame, bg="#f5f5f5")
        self.services_container.grid(row=row_idx, column=0, columnspan=2, sticky="ew", pady=(5, 5)); row_idx += 1
        
        self.service_rows = []

        def add_service_row():
            row_frame = tk.Frame(self.services_container, bg="#f5f5f5")
            row_frame.pack(fill="x", pady=2)

            service_var = tk.StringVar()
            combo = ttk.Combobox(
                row_frame, textvariable=service_var,
                values=SERVICES, state="readonly",
                font=("Helvetica", 11), width=18
            )
            combo.pack(side="left", padx=(0, 10))

            price_var = tk.StringVar()
            price_entry = tk.Entry(
                row_frame, textvariable=price_var,
                font=("Helvetica", 11),
                relief="solid", bd=1, bg="white", width=12
            )
            price_entry.pack(side="left", padx=(0, 10))
            price_var.trace_add("write", self._recalc)
            
            service_data = {
                "frame": row_frame,
                "service": service_var,
                "price": price_var
            }
            
            if len(self.service_rows) > 0:
                btn_remove = tk.Button(
                    row_frame, text="✖", font=("Helvetica", 10),
                    fg="red", bg="#f5f5f5", relief="flat", cursor="hand2",
                    command=lambda: remove_service_row(row_frame, service_data)
                )
                btn_remove.pack(side="left")

            self.service_rows.append(service_data)

        def remove_service_row(row_frame, service_data):
            row_frame.destroy()
            self.service_rows.remove(service_data)
            self._recalc()

        add_service_row()
        
        btn_add = tk.Button(
            form_frame, text="+ Добавить услугу", font=("Helvetica", 10),
            bg="#e0e0e0", fg="#333", relief="flat", cursor="hand2",
            command=add_service_row
        )
        btn_add.grid(row=row_idx, column=0, columnspan=2, sticky="w", pady=(0, 10)); row_idx += 1

        add_label_entry(row_idx,  "Материал (₸)", "material"); row_idx += 1
        add_label_entry(row_idx,  "Сумма рабочего", "worker",  readonly=True); row_idx += 1
        add_label_entry(row_idx,  "Расход (₸)",   "rashod"); row_idx += 1
        add_label_entry(row_idx,  "Касса (₸)",    "kassa",    readonly=True); row_idx += 1
        add_label_entry(row_idx,  "Примечание",   "note"); row_idx += 1

        # Live calculation bindings
        self.entries["material"].trace_add("write", self._recalc)
        self.entries["rashod"].trace_add("write", self._recalc)

        # Divider
        tk.Frame(self.root, height=1, bg="#ddd").pack(fill="x", padx=24)

        # Save button
        btn_frame = tk.Frame(self.root, bg="#f5f5f5", pady=14)
        btn_frame.pack()
        tk.Button(
            btn_frame,
            text="💾  Сохранить",
            font=("Helvetica", 12, "bold"),
            bg="#1a1a2e", fg="white",
            activebackground="#16213e",
            activeforeground="white",
            relief="flat", padx=28, pady=8,
            cursor="hand2",
            command=self._save
        ).pack()

        tk.Label(
            self.root,
            text="* — обязательные поля",
            font=("Helvetica", 9), fg="#999", bg="#f5f5f5"
        ).pack(pady=(0, 8))

    def _clear_hint(self, event, widget, hint):
        if widget.get() == hint:
            widget.delete(0, tk.END)
            widget.config(fg="black")

    def _restore_hint(self, event, widget, hint):
        if widget.get() == "":
            widget.insert(0, hint)
            widget.config(fg="grey")

    def _safe_float(self, key):
        try:
            val = self.entries[key].get().strip()
            if val in ("", "87XXXXXXXXX"):
                return 0.0
            return float(val)
        except ValueError:
            return 0.0

    def _recalc(self, *args):
        total_price = 0.0
        for row in self.service_rows:
            val = row["price"].get().strip()
            if val:
                try:
                    total_price += float(val)
                except ValueError:
                    pass

        material = self._safe_float("material")
        rashod = self._safe_float("rashod")
        self.entries["worker"].set(str(round(total_price - material, 2)))
        self.entries["kassa"].set(str(round(total_price - rashod, 2)))

    def _get_next_number(self):
        if not os.path.exists(EXCEL_FILE):
            return 1
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        return ws.max_row  # row 1 = header, so max_row - 1 data rows + 1 = max_row

    def _ensure_excel(self):
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(HEADERS)
            # Style header
            from openpyxl.styles import Font, PatternFill, Alignment
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1a1a2e")
                cell.alignment = Alignment(horizontal="center")
            wb.save(EXCEL_FILE)

    def _save(self):
        marka   = self.entries["marka"].get().strip()
        gnomer  = self.entries["gnomer"].get().strip()

        # Validation
        missing = []
        if not marka:   missing.append("Марка")
        if not gnomer:  missing.append("г.номер")
        
        valid_services = []
        total_price = 0.0
        
        for row in self.service_rows:
            srv = row["service"].get().strip()
            prc_s = row["price"].get().strip()
            
            if srv and prc_s:
                try:
                    prc = float(prc_s)
                    valid_services.append(f"{srv} ({prc})")
                    total_price += prc
                except ValueError:
                    messagebox.showerror("Ошибка", f"❌ Цена для услуги '{srv}' должна быть числом")
                    return
            elif srv and not prc_s:
                missing.append("Цена")
            elif not srv and prc_s:
                missing.append("Услуга")

        if not valid_services and not missing:
            missing.append("Услуга")
            missing.append("Цена")

        if missing:
            # Deduplicate missing fields
            missing = list(dict.fromkeys(missing))
            messagebox.showerror(
                "Ошибка",
                f"❌ Заполните обязательные поля: {', '.join(missing)}"
            )
            return

        service_str = ", ".join(valid_services)

        owner    = self.entries["owner"].get().strip()
        if owner == "87XXXXXXXXX":
            owner = ""
        material = self._safe_float("material")
        rashod   = self._safe_float("rashod")
        worker   = round(total_price - material, 2)
        kassa    = round(total_price - rashod, 2)
        note     = self.entries["note"].get().strip()

        self._ensure_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        num = ws.max_row  # header is row 1, so next num = max_row
        row_num = f"#{num}"
        ws.append([row_num, marka, gnomer, owner, service_str,
                   total_price, material, worker, rashod, kassa, note])
        wb.save(EXCEL_FILE)

        messagebox.showinfo("Сохранено", f"✅ Запись {row_num} сохранена")
        self._clear_form()

    def _clear_form(self):
        skip_readonly = {"worker", "kassa"}
        hints = {"owner": "87XXXXXXXXX"}
        for key, var in self.entries.items():
            if key in skip_readonly:
                var.set("0")
            else:
                var.set("")
        
        # Clear service rows and leave only one
        for row in self.service_rows[1:]:
            row["frame"].destroy()
        self.service_rows = [self.service_rows[0]]
        self.service_rows[0]["service"].set("")
        self.service_rows[0]["price"].set("")

        # Restore hints visually
        for widget in self.root.winfo_children():
            self._restore_hints_recursive(widget, hints)

    def _restore_hints_recursive(self, widget, hints):
        if isinstance(widget, tk.Entry):
            pass  # hints handled by FocusOut bindings
        for child in widget.winfo_children():
            self._restore_hints_recursive(child, hints)


def main():
    root = tk.Tk()
    app = DetailingApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
